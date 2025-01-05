#!/usr/bin/env python
# -*- coding: UTF-8 -*-

# pip install --upgrade numpy
# pip install --upgrade pandas openpyxl
# pip install requests 

# 使用: python get_entries.py

import os
import configparser
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# 每次失败后等待2秒钟，并最多重试3次
def get_data_from_url(url, retries=3, delay=2):
    for i in range(retries):
        try:
            response = requests.get(url)
            response.raise_for_status()  # 如果响应状态码不是200，将引发HTTPError异常
            return response.json()
        except requests.exceptions.HTTPError as errh:
            print(f"HTTP Error: {errh}")
        except requests.exceptions.ConnectionError as errc:
            print(f"Error Connecting: {errc}")
        except requests.exceptions.Timeout as errt:
            print(f"Timeout Error: {errt}")
        except requests.exceptions.RequestException as err:
            print(f"Error: {err}")
        
        print(f"Retrying in {delay} seconds...")
        time.sleep(delay)
    
    return None  # 如果所有重试都失败，则返回 None

# 返回键是 index 值是词条的洗孔字典
def get_index_wash_entries(seed, start_index, count=100):
    index_wash_entries = {}
    for num in range(0, count // 50):  # 每个请求获取50个条目
        url = f"https://hbrapi.fuyumi.xyz/api/ChangeAbility?_seed={seed}&_index={start_index}"
        data = get_data_from_url(url)
        index_wash_entries.update({str(int(start_index) + i + 1): data[str(i)] for i in range(50)})
        start_index = str(int(start_index) + 50)
    return index_wash_entries

# 返回键是 index 值是词条的装备字典
def get_index_equipments(seed, start_index, count=100):
    index_equipments = {}
    for num in range(0, count // 50):  # 每个请求获取50个条目
        url = f"https://hbrapi.fuyumi.xyz/api/RandomMainAbility?_seed={seed}&_index={start_index}"
        data = get_data_from_url(url)
        # index_equipments.update({str(int(start_index) + i + 1): data[str(i)] for i in range(50)})
        index_equipments.update({str(int(start_index) + i + 1): data[str(i)].split('/') for i in range(50)})
        start_index = str(int(start_index) + 50)
    return index_equipments

def print_dir(dir_data):
    for key, value in dir_data.items():
        print(f"{key}: {value}")

def save_index_equipments_to_file(index_equipments):
    # 将字典转换为 DataFrame
    df = pd.DataFrame.from_dict(index_equipments, orient='index', 
        columns=['第一词条', 'DP', '智慧', '通常攻击攻击力', '体力', '精神', '初始SP'])
    df.reset_index(inplace=True)
    df.rename(columns={'index': '索引'}, inplace=True)

    # 定义黄色填充样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    excel_file_path = './index_equipments.xlsx'
    # 使用 ExcelWriter 来保存并应用样式
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        
        # 查找 "DP +1200" 并填充整行为黄色
        for idx, row in df.iterrows():
            if row['DP'] == "DP +1200":
                for col in range(1, len(df.columns) + 1):  # +1 因为 openpyxl 是 1-indexed
                    worksheet.cell(row=idx + 2, column=col).fill = yellow_fill

        # 设置每列的宽度
        column_widths = [10, 14, 12, 12, 22, 12, 12, 20]
        for i, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

def save_index_wash_entries_to_file(index_wash_entries):
    # 将字典转换为 DataFrame
    df = pd.DataFrame.from_dict(index_wash_entries, orient='index', columns=['词条'])

    df.reset_index(inplace=True)
    df.rename(columns={'index': '索引'}, inplace=True)

    # 定义黄色填充样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    excel_file_path = './index_wash_entries.xlsx'
    # 使用 ExcelWriter 来保存并应用样式
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        for idx, row in df.iterrows():
            if '+3' in str(row['词条']) and ('+30' not in str(row['词条'])):
                for col in range(1, len(df.columns) + 1):  # +1 因为 openpyxl 是 1-indexed
                    worksheet.cell(row=idx + 2, column=col).fill = yellow_fill  # +2因为标题行

        # 设置每列的宽度
        column_widths = [10, 14]
        for i, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[get_column_letter(i)].width = width


def check_empty_values(
    ChangeAbility_seed, ChangeAbility_index, 
    RandomMainAbility_seed, RandomMainAbility_index, 
    DataCount):
    
    empty_variables = []
    if not ChangeAbility_seed:
        empty_variables.append('ChangeAbility_seed')
    if not ChangeAbility_index:
        empty_variables.append('ChangeAbility_index')
    if not RandomMainAbility_seed:
        empty_variables.append('RandomMainAbility_seed')
    if not RandomMainAbility_index:
        empty_variables.append('RandomMainAbility_index')
    if DataCount is None:
        empty_variables.append('DataCount')

    empty_variables_message = "以下变量为空: " + ", ".join(empty_variables) 
    if empty_variables:
        print(empty_variables_message)
        return True
    else:
        return False

def main():
    try:
        config = configparser.ConfigParser()
        config.read('config.ini')

        # 将配置文件中的值分配给变量
        # 洗词条
        ChangeAbility_seed = config.get('ChangeAbility', 'ChangeAbility_seed', fallback="")
        ChangeAbility_index = config.get('ChangeAbility', 'ChangeAbility_index', fallback="")
        # 打装备
        RandomMainAbility_seed = config.get('RandomMainAbility', 'RandomMainAbility_seed', fallback="")
        RandomMainAbility_index = config.get('RandomMainAbility', 'RandomMainAbility_index', fallback="")
        # 获取数据条目数
        DataCount = config.getint('Count', 'DataCount')

        if check_empty_values(ChangeAbility_seed, ChangeAbility_index, 
            RandomMainAbility_seed, RandomMainAbility_index, 
            DataCount):
            print("[-] 请修改配置文件 config.ini")
        else:
            print("洗词条:")
            index_wash_entries = get_index_wash_entries(
                ChangeAbility_seed, 
                ChangeAbility_index, 
                count=DataCount
            )
            print_dir(index_wash_entries)
            # 保存为 Excel 文件
            save_index_wash_entries_to_file(index_wash_entries)

            print("打装备:")
            index_equipments = get_index_equipments(
                RandomMainAbility_seed, 
                RandomMainAbility_index, 
                count=DataCount
            )
            print_dir(index_equipments)
            # 保存为 Excel 文件
            save_index_equipments_to_file(index_equipments)

        input()

    except Exception as e:
        print(f"[-] {e}")

if __name__ == "__main__":
    main()
