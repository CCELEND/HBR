#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

import tkinter as tk
import os
import time
import configparser
import pandas as pd
import openpyxl
from tkinter import scrolledtext, Menu, messagebox
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import math
import datetime
import threading
from collections import OrderedDict


from proc import parallel_process_indexes

index_wash_entries = {}
index_wash_entries_lock = threading.Lock()  # 创建一个锁
index_equipments = {}
index_equipments_lock = threading.Lock()  # 创建一个锁

# 右键复制
def copy_text(event, text_widget):
    try:
        # 获取选中的文本
        selected_text = text_widget.get("sel.first", "sel.last")
        if selected_text:
            text_widget.clipboard_clear()
            text_widget.clipboard_append(selected_text)
    except tk.TclError:
        pass  # 如果没有选中文本，忽略错误

# 右键粘贴
def paste_text(event, text_widget):
    try:
        text_widget.insert(tk.INSERT, text_widget.clipboard_get())
    except tk.TclError:
        pass  # 如果剪贴板为空，忽略错误

# 右键剪切
def cut_text(event, text_widget):
    try:
        selected_text = text_widget.get("sel.first", "sel.last")
        if selected_text:
            text_widget.clipboard_clear()
            text_widget.clipboard_append(selected_text)
            text_widget.delete("sel.first", "sel.last")
    except tk.TclError:
        pass

# 右键菜单
def show_context_menu(event, text_widget):
    # 创建上下文菜单
    context_menu = Menu(text_widget, tearoff=0)
    context_menu.add_command(label="复制", command=lambda e=event: copy_text(e, text_widget))
    context_menu.add_command(label="粘贴", command=lambda e=event: paste_text(e, text_widget))
    context_menu.add_command(label="剪切", command=lambda e=event: cut_text(e, text_widget))
    # 在鼠标右键点击的位置显示菜单
    context_menu.tk_popup(event.x_root, event.y_root)
    context_menu.grab_release()

def clear_entries():
    index_wash_entries.clear()
    index_equipments.clear()

# 清空输入输出框及字典
def clear_text(*text_widgets):
    for text_widget in text_widgets:
        if text_widget.cget('state') == tk.DISABLED:
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            text_widget.config(state=tk.DISABLED)
        else:
            text_widget.delete("1.0", tk.END)

    clear_entries()

# 编辑文本框
def edit_text(text_widget, data):
    if text_widget.cget('state') == tk.DISABLED:
        text_widget.config(state=tk.NORMAL)
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, data)
        text_widget.config(state=tk.DISABLED)
    else:
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, data)

def print_dir(dir_data):
    data = ""
    for key, value in dir_data.items():
        data += f"{key}: {value}\n"

    edit_text(output_text, data)

def check_wash_entries(
    ChangeAbility_seed, ChangeAbility_index, 
    DataCount):
    
    empty_variables = []
    if not ChangeAbility_seed:
        empty_variables.append('ChangeAbility_seed')
    if not ChangeAbility_index:
        empty_variables.append('ChangeAbility_index')
    if not DataCount:
        empty_variables.append('DataCount')

    empty_variables_message = "以下变量为空: " + ", ".join(empty_variables) + "\n请修改配置文件 config.ini"
    if empty_variables:
        messagebox.showerror("错误", empty_variables_message)
        return True
    else:
        return False

def check_equipments(
    RandomMainAbility_seed, RandomMainAbility_index, 
    DataCount):
    
    empty_variables = []
    if not RandomMainAbility_seed:
        empty_variables.append('RandomMainAbility_seed')
    if not RandomMainAbility_index:
        empty_variables.append('RandomMainAbility_index')
    if not DataCount:
        empty_variables.append('DataCount')

    empty_variables_message = "以下变量为空: " + ", ".join(empty_variables) + "\n请修改配置文件 config.ini"
    if empty_variables:
        messagebox.showerror("错误", empty_variables_message)
        return True
    else:
        return False


# 获取洗孔的词条字典
def get_index_wash_entries():

    if not os.path.exists("config.ini"):
        messagebox.showerror("错误", "配置文件 config.ini 不存在")
        return

    config = configparser.ConfigParser()
    config.read('config.ini')

    # 将配置文件中的值分配给变量
    # 洗词条
    ChangeAbility_seed = config.get('ChangeAbility', 'ChangeAbility_seed', fallback="")
    ChangeAbility_index = config.get('ChangeAbility', 'ChangeAbility_index', fallback="")
    # 获取数据条目数
    DataCount = config.get('Count', 'DataCount', fallback="")

    if check_wash_entries(ChangeAbility_seed, ChangeAbility_index, DataCount):
        return

    clear_entries()

    seed = int(ChangeAbility_seed)
    start_index = int(ChangeAbility_index)
    end_index = start_index + int(DataCount)

    # starttime = datetime.datetime.now()
    # for cur_index in range(start_index, end_index):
    #     real = get_random_value(seed, cur_index)
    #     wash_entry_str = get_property(real, wash_entry)
    #     index_wash_entries.update({ str(cur_index):[wash_entry_str, str(real)]})
    # endtime = datetime.datetime.now()

    starttime = datetime.datetime.now()
    global index_wash_entries
    temp_index_wash_entries = parallel_process_indexes(
        fun=0,
        seed=seed,
        start_index=start_index,
        end_index=end_index,
        chunk_size=10,
        max_workers=max(4, os.cpu_count())
    )
    index_wash_entries = OrderedDict(sorted(temp_index_wash_entries.items()))
    endtime = datetime.datetime.now()

    print("use times {0:.2f}s".format((endtime - starttime).total_seconds()))
    print_dir(index_wash_entries)

# 获取装备的词条字典
def get_index_equipments():

    if not os.path.exists("config.ini"):
        messagebox.showerror("错误", "配置文件 config.ini 不存在！")
        return

    config = configparser.ConfigParser()
    config.read('config.ini')

    # 将配置文件中的值分配给变量
    # 打装备
    RandomMainAbility_seed = config.get('RandomMainAbility', 'RandomMainAbility_seed', fallback="")
    RandomMainAbility_index = config.get('RandomMainAbility', 'RandomMainAbility_index', fallback="")
    # 获取数据条目数
    DataCount = config.get('Count', 'DataCount', fallback="")

    if check_equipments(RandomMainAbility_seed, RandomMainAbility_index, DataCount):
        return

    clear_entries()

    seed = int(RandomMainAbility_seed)
    start_index = int(RandomMainAbility_index)
    end_index = start_index + int(DataCount)

    # for cur_index in range(start_index, end_index):
    #     real = get_random_value(seed, cur_index)
    #     spct_str = get_property(real, spct)
    #     career_entry_str = get_property(real, career_entry)
    #     charm_entry_str = get_property(real, charm_entry)
    #     spct_list = spct_str.split()
    #     index_equipments.update({ str(cur_index):spct_list+[career_entry_str+"的初始SP+3", charm_entry_str, str(real)]})

    starttime = datetime.datetime.now()
    global index_equipments
    temp_index_equipments = parallel_process_indexes(
        fun=1,
        seed=seed,
        start_index=start_index,
        end_index=end_index,
        chunk_size=10,
        max_workers=max(4, os.cpu_count())
    )
    index_equipments = OrderedDict(sorted(temp_index_equipments.items()))
    endtime = datetime.datetime.now()

    print("use times {0:.2f}s".format((endtime - starttime).total_seconds()))
    print_dir(index_equipments)

def save_to_file():
    if index_wash_entries:
        save_index_wash_entries_to_file(index_wash_entries)
    elif index_equipments:
        save_index_equipments_to_file(index_equipments)
    else:
        messagebox.showinfo("提示", "无数据，请获取词条")


def save_index_equipments_to_file(index_equipments):

    # 将字典转换为 DataFrame
    df = pd.DataFrame.from_dict(index_equipments, orient='index', 
        columns=['第一词条', 'DP', '智慧', '通常攻击攻击力', '体力', '精神', '初始SP', '吊饰' ,'真实随机值'])
    df.reset_index(inplace=True)
    df.rename(columns={'index': '索引'}, inplace=True)

    # 定义黄色填充样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    excel_file_path = './index_equipments.xlsx'
    try:
        # 使用 ExcelWriter 来保存并应用样式
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # 查找 "DP +1200" 并填充整行为黄色
            for idx, row in df.iterrows():
                if row['DP'] == "DP+1200":
                    for col in range(1, len(df.columns) + 1):  # +1 因为 openpyxl 是 1-indexed
                        worksheet.cell(row=idx + 2, column=col).fill = yellow_fill

            # 设置每列的宽度
            column_widths = [10, 14, 12, 12, 22, 12, 12, 20, 12, 14]
            for i, width in enumerate(column_widths, start=1):
                worksheet.column_dimensions[get_column_letter(i)].width = width

        messagebox.showinfo("提示", "装备词条数据已保存至: \nindex_equipments.xlsx")
    except Exception as e:
        messagebox.showerror("错误", f"{e}\n请关闭打开的 index_equipments.xlsx 并重试")

def save_index_wash_entries_to_file(index_wash_entries):
    # 将字典转换为 DataFrame
    df = pd.DataFrame.from_dict(index_wash_entries, orient='index', columns=['词条','真实随机值'])

    df.reset_index(inplace=True)
    df.rename(columns={'index': '索引'}, inplace=True)

    # 定义黄色填充样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    excel_file_path = './index_wash_entries.xlsx'
    try:
        # 使用 ExcelWriter 来保存并应用样式
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']

            for idx, row in df.iterrows():
                if '+3' in str(row['词条']) and ('+30' not in str(row['词条'])):
                    for col in range(1, len(df.columns) + 1):  # +1 因为 openpyxl 是 1-indexed
                        worksheet.cell(row=idx + 2, column=col).fill = yellow_fill  # +2因为标题行

            # 设置每列的宽度
            column_widths = [10, 14, 14]
            for i, width in enumerate(column_widths, start=1):
                worksheet.column_dimensions[get_column_letter(i)].width = width

        messagebox.showinfo("提示", "洗孔词条数据已保存至: \nindex_wash_entries.xlsx")
    except Exception as e:
        messagebox.showerror("错误", f"{e}\n请关闭打开的 index_wash_entries.xlsx 并重试")

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("700x405")
    root.title("词条获取")

    # 配置主窗口的列和行的伸展
    root.grid_rowconfigure(0, weight=1)     # get_entries_frame 行随着窗口调整大小
    root.grid_columnconfigure(0, weight=1)  # 第0列会随着窗口调整大小
    root.grid_columnconfigure(1, weight=1)  # 第1列会随着窗口调整大小
    root.grid_columnconfigure(2, weight=1)  # 第2列会随着窗口调整大小
    root.grid_columnconfigure(3, weight=1)  # 第3列会随着窗口调整大小


    # 创建一个 Frame
    get_entries_frame = tk.Frame(root)
    get_entries_frame.grid(row=0, column=0, columnspan=4, padx=0, pady=(10,20), sticky="nsew")
    # 配置框架的行列的伸展框架的伸展
    get_entries_frame.grid_rowconfigure(0, weight=1)      
    get_entries_frame.grid_columnconfigure(0, weight=1)  # 使输出框占满整列
    get_entries_frame.grid_columnconfigure(1, weight=1)  # 使输出框占满整列
    get_entries_frame.grid_columnconfigure(2, weight=1)  # 使输出框占满整列
    get_entries_frame.grid_columnconfigure(3, weight=1)  # 使输出框占满整列
      

    # 输出框
    output_text = scrolledtext.ScrolledText(get_entries_frame, 
        wrap=tk.WORD, width=50, height=20)
    output_text.grid(row=0, column=0, columnspan=3, padx=(10,0), pady=0, sticky="nsew")
    # 绑定鼠标右键点击事件到上下文菜单
    output_text.bind("<Button-3>", lambda event, tw=output_text: show_context_menu(event, tw))

    # 按钮框架
    buttons_frame = tk.Frame(get_entries_frame)
    buttons_frame.grid(row=0, column=3, padx=(0,10), pady=0, sticky="nsew")
    buttons_frame.grid_rowconfigure(0, weight=1)
    buttons_frame.grid_rowconfigure(1, weight=1)
    buttons_frame.grid_rowconfigure(2, weight=1)
    buttons_frame.grid_rowconfigure(3, weight=1)
    buttons_frame.grid_columnconfigure(0, weight=1)  # 占满整列



    # 洗孔词条按钮
    get_index_equipments_button = tk.Button(buttons_frame, 
        width=20, text="获取洗孔词条", command=lambda: get_index_wash_entries())
    get_index_equipments_button.grid(row=0, column=0, padx=(0,10), pady=(10,0))
    # 装备词条按钮
    get_index_equipments_button = tk.Button(buttons_frame, 
        width=20, text="获取装备词条", command=lambda: get_index_equipments())
    get_index_equipments_button.grid(row=1, column=0, padx=(0,10), pady=(10,0))


    # 创建清空按钮
    clear_button = tk.Button(buttons_frame, 
        width=20, text="清空", command=lambda: clear_text(output_text))
    clear_button.grid(row=2, column=0, padx=(0,10), pady=(10,0))

    # 创建保存 Excel 文件按钮
    save_file_button = tk.Button(buttons_frame, 
        width=20, text="保存为 Excel 文件", command=save_to_file)
    save_file_button.grid(row=3, column=0, padx=(0,10), pady=(10,0))

    root.mainloop()
